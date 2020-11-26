"""
    build an excel file from data
"""
from copy import deepcopy

import click

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments.comments import Comment
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill, GradientFill, Font, Color, colors, Alignment, Border, Side

from simple_loggers import SimpleLogger


class ExcelBuilder(object):
    logger = SimpleLogger('ExcelBuilder')
    def __init__(self):
        self.wb = openpyxl.Workbook(write_only=False)
        self.wb.remove(self.wb.active)  # remove default 'Sheet'

    def create_sheet(self, title=None):
        """
            require: len(title) <= 31
        """
        if title and len(title) > 31:
            self.logger.warning('title is to long, limit 31 characters')
            title = title[:31]
        self.logger.info('create sheet: {}'.format(title))
        self.ws = self.wb.create_sheet(title)

    def add_title(self, titles, **style):
        self.logger.debug('>>> add title')
        self.ws.append(titles)
        self.set_row_style(**style)

    def add_rows(self, rows, color_list=None, **style):
        self.logger.info('>>> add rows')

        self.logger.debug('color_list: {}, style: {}'.format(color_list, list(style.keys())))

        for n, row in enumerate(rows, 1):
            self.ws.append(row)
            if not (color_list or style):
                continue

            if color_list:
                color = color_list[n % len(color_list)]
                style['PatternFill'] = PatternFill(start_color=color, end_color=color, fill_type='solid')
            self.set_row_style(**style)

    def set_row_style(self, **style):
        if not style:
            return

        for cell in self.ws[self.ws.max_row]:
            if style.get('font'):
                cell.font = style['font']

            if style.get('alignment'):
                cell.alignment = style['alignment']

            if style.get('border'):
                cell.border = style['border']

            if style.get('PatternFill'):
                # print(cell.coordinate, cell.value, style['PatternFill'].start_color.value)
                cell.fill = style['PatternFill']
            elif style.get('GradientFill'):
                cell.fill = style['GradientFill']

    def set_dimensions_style(self, height=None, width=None, **style):
        self.logger.info('>>> set dimensions style')
        height = height or (style and style.get('height'))
        width = width or (style and style.get('width'))
        if height:
            for row in range(1, self.ws.max_row + 1):
                r = self.ws.row_dimensions[row]
                r.height = height
            
        if width:
            for column in range(1, self.ws.max_column + 1):
                c= self.ws.column_dimensions[get_column_letter(column)]
                c.width = width

    def freeze_panes(self, first_column=None, first_row=True, first_column_and_row=None, coordinate=None):
        if not coordinate:
            if first_column:
                coordinate = 'B1'
            elif first_row:
                coordinate = 'A2'
            elif first_column_and_row:
                coordinate = 'B2'
        if coordinate:
            self.logger.info('>>> freeze: {}'.format(coordinate))
            self.ws.freeze_panes = coordinate

    def add_comment(self, coordinate, *args, **kwargs):
        self.logger.info('>>> add comment')
        self.ws[coordinate].comment = Comment(*args, **kwargs)

    def add_hyperlink(self, coordinate, hyperlink):
        """
            # method1:
            >>> value = '=HYPERLINK("http://www.baidu.com", "baidu")'
            >>> sheet.cell(row, colum, value=value)

            # method2:
            >>> _ = sheet.cell(row, column, value='baidu')
            >>> _.hyperlink = 'http://www.baidu.com'

            # method3:
            >>> sheet['B3'].hyperlink = Hyperlink(ref='', target='http://www.baidu.com', tooltip='baidu')
        """
        self.logger.info('>>> add hyperlink')
        self.ws[coordinate].hyperlink = hyperlink

    def merge_cells(self, *args, **kwargs):
        """
            range_string=None
            start_row=None
            start_column=None
            end_row=None
            end_column=None
        """
        self.logger.info('>>> merge cells')
        self.ws.merge_cells(*args, **kwargs)

    def auto_filter(self):
        """
            automatical filter for first row
        """
        self.ws.auto_filter.ref = self.ws.dimensions

    def save(self, outfile):
        self.wb.save(outfile)
        self.logger.warning('save file: {}'.format(outfile))


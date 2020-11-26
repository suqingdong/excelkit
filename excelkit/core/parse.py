"""
    parse excel file
"""
import sys
import json
import datetime
from collections import OrderedDict

import click
import openpyxl

from simple_loggers import SimpleLogger
from excelkit.util.reformat import Formatter


class ExcelParser(object):
    logger = SimpleLogger('ExcelParser')
    def __init__(self):
        pass

    def parse(self, filename, data_only=False, read_only=False, sheet_idx=None, choose_one=False, skip=None, limit=None, **kwargs):
        """
            data_only=True: get the value instead of formula when data_type is 'f'
            read_only=True: to deal with large file, some attributes might lost
        """
        wb = openpyxl.load_workbook(filename, data_only=data_only, read_only=read_only)

        sheets = wb.worksheets
        if len(wb.sheetnames) > 1:
            if sheet_idx is not None:
                sheets = [wb.worksheets[sheet_idx]]
            elif choose_one:
                sheets = self.choose_sheet(wb)

        return self.get_data(sheets, skip=skip, limit=limit)

    def choose_sheet(self, workbook):
        context = dict(enumerate(workbook.sheetnames))
        click.echo('{}'.format(json.dumps(context, ensure_ascii=False)), err=True)

        while True:
            idxes = click.prompt('please choose one or more sheets, separate by comma', err=True)
            for idx in idxes.split(','):
                if int(idx) not in context:
                    self.logger.warning('bad choice, choose from: {}'.format(list(context.keys())))
                    continue
            sheets = [workbook.worksheets[int(idx)] for idx in idxes]
            return sheets

    def get_data(self, worksheets, skip=None, limit=None, fillna=''):
        data = OrderedDict()
        for ws in worksheets:
            data[ws.title] = []
            for n, row in enumerate(ws.rows):
                if skip and n < skip:
                    continue
                if limit and len(data[ws.title]) > limit:
                    break

                line = [
                    cell.value.strftime('%Y-%m-%d')
                    if cell.data_type == 'd'
                    else fillna if cell.value is None else cell.value
                    for cell in row
                ]
                data[ws.title].append(line)

        return data

    def export(self, data, outfile=None, fmt='table', indent=None, sep='\t', header=True, index=True, color=None, pager=False):
        """"export data

        parameters
            data: data return by get_data method
            outfile: output file, default stdout
            fmt: 'table', 'html', 'tsv' or 'json'
            indent: for json fmt export
            sep: for tsv fmt export
        """
        out = open(outfile, 'w') if outfile else sys.stdout
        with out:
            for sheet, rows in data.items():
                click.secho('>>> {}'.format(sheet), err=True, fg='yellow')
                fd = Formatter(rows, header=header)
                if fmt == 'table':
                    res = fd.to_table(index=index).get_string()
                elif fmt == 'html':
                    res = fd.to_table(index=index).get_html_string()
                elif fmt == 'json':
                    res = fd.to_json(indent=indent)
                elif fmt == 'tsv':
                    res = fd.to_tsv(sep=sep)
                else:
                    exit('bad format, choose from table, html, json, tsv')

                if color:
                    res = click.style(res, fg=color)

                if pager:
                    click.echo_via_pager(res, color=False)
                else:
                    out.write(res + '\n')


def parse_text(file_hdl, sep='\t', comment=None):

    with file_hdl:
        for line in file_hdl:
            if comment and line.startswith(comment):
                continue
            linelist = line.strip().split(sep)
            yield linelist

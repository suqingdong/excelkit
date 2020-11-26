"""
    concat multiple excel files to one
"""
import openpyxl

from simple_loggers import SimpleLogger


class ExcelConcat(object):
    logger = SimpleLogger('ExcelConcat')
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # remove default 'Sheet'

    def concat(self, infiles, keep_fmt=False, keep_size=True, merge_cells=True, keep_other=True):
        for infile in infiles:
            self.logger.debug('loading file: {}'.format(infile))
            in_book = openpyxl.load_workbook(infile)
            for sheetname in in_book.sheetnames:
                sheet = in_book[sheetname]
                self.logger.debug('copy sheet: {} [{} rows, {} columns]'.format(sheetname, sheet.max_row, sheet.max_column))
                ws = self.wb.create_sheet(sheetname)
                for row in sheet.rows:
                    for cell in row:
                        ws[cell.coordinate] = cell.value
                        if keep_fmt and cell.has_style:    # might be slow for big file
                            self.copy_format(cell, ws[cell.coordinate])
                if keep_size:
                    self.copy_size(sheet, ws)
                if merge_cells:
                    self.merge_cells(sheet, ws)
                    
                if keep_other:
                    self.copy_other(sheet, ws)

    def copy_format(self, origin_cell, target_cell):
        """
            copy style for each cell
        """
        fmt_list = (
            'alignment', 'font', 'fill', 'border', 'comment',
            'hyperlink', 'data_type', 'number_format'
        )
        for fmt in fmt_list:
            value = getattr(origin_cell, fmt)
            if not value:
                continue
            if isinstance(value, openpyxl.styles.proxy.StyleProxy):
                value = value.copy()
            setattr(target_cell, fmt, value)

    def copy_size(self, origin_sheet, target_sheet):
        """
            copy width for columns and height for rows
        """
        self.logger.debug('copy height and width for sheet: {}'.format(origin_sheet.title))
        for column in range(1, origin_sheet.max_column + 1):
            letter = openpyxl.utils.get_column_letter(column)
            width = origin_sheet.column_dimensions[letter].width
            target_sheet.column_dimensions[letter].width = origin_sheet.column_dimensions[letter].width
        
        for row in range(1, origin_sheet.max_row + 1):
            target_sheet.row_dimensions[row].height = origin_sheet.row_dimensions[row].height

    def merge_cells(self, origin_sheet, target_sheet):
        """
            copy merged cells
        """
        self.logger.debug('merge cells for sheet: {}'.format(origin_sheet.title))
        for ranges in origin_sheet.merged_cell_ranges:
            target_sheet.merge_cells(ranges.coord)

    def copy_image(self, origin_sheet, target_sheet):
        self.logger.debug('copy images for sheet: {}'.format(origin_sheet.title))
        for im in origin_sheet._images:
            target_sheet.add_image(im)

    def copy_other(self, origin_sheet, target_sheet):
        for other in ('image', 'table', 'chart', 'pivot'):
            items = getattr(origin_sheet, '_{}s'.format(other))

            if other == 'table':  # dict for table
                items = items.values()

            if items:
                self.logger.debug('copy {} for sheet: {}'.format(other, origin_sheet.title))
                for item in items:
                    getattr(target_sheet, 'add_{}'.format(other))(item)

        if origin_sheet.data_validations.dataValidation:
            self.logger.debug('copy data_validations for sheet: {}'.format(origin_sheet.title))
            for data_validation in origin_sheet.data_validations.dataValidation:
                target_sheet.add_data_validation(data_validation)

    def save(self, outfile):
        self.wb.save(outfile)
        self.logger.info('save file: {}'.format(outfile))


if __name__ == '__main__':
    concat = ExcelConcat()
    concat.concat(['demo/demo.xlsx', 'demo/info.B35S7.xlsx'], keep_fmt=True, keep_size=True, merge_cells=True)
    concat.save('demo/out.xlsx')

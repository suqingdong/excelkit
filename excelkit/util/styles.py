from openpyxl.styles import PatternFill, GradientFill, Font, Color, colors, Alignment, Border, Side



WHITE_FILL = PatternFill(start_color=colors.WHITE, end_color=colors.WHITE, fill_type='solid')
BLACK_FILL = PatternFill(start_color=colors.BLACK, end_color=colors.BLACK, fill_type='solid')
ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


WHITE_ON_BLACK = {
    'width': None,
    'height': None,
    'alignment': ALIGNMENT_CENTER,
    'font': Font(color=colors.WHITE, name=None, size=None, bold=True, italic=False),
    'PatternFill': BLACK_FILL,
    'GradientFill': GradientFill(type='linear', stop=())
}


HEAD_STYLES = {
    'white_on_black': WHITE_ON_BLACK,
}

BODY_STYLES = {
    'cyan_green': ('B3FFB3', 'B3FFFF'),
    'white_grey': ('FFFFFF', 'CCCCCC'),
}

THREE_LINE_TABLE = {
    'body': {
        'PatternFill': WHITE_FILL,
        'alignment': ALIGNMENT_CENTER,
        'font': Font(name='Times New Roman')
    },
    'head': {
        'border': Border(top=Side(style='thin'), bottom=Side(style='thin')),
        'PatternFill': WHITE_FILL,
        'alignment': ALIGNMENT_CENTER,
        'font': Font(name='Times New Roman', bold=True)
    },
    'foot': {
        'border': Border(bottom=Side(style='thin')),
        'PatternFill': WHITE_FILL,
        'alignment': ALIGNMENT_CENTER,
    }
}
